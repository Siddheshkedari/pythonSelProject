import openpyxl


class HomePageData:

    test_Homepage_Data = [{"Name" : "Siddhesh","email" : "sp@gmail.com","gender":"Male"}, {"Name" : "Manisha","email" : "sprm@gmail.com","gender":"Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\91966\\Documents\\pythonDemo.xlsx")  # path = C:\Users\91966\Documents
        sheet = book.active  # returns the active sheet
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]